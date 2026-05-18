#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导出聚水潭订单明细Excel - 修正版"""

import sys
import os
import datetime as dt
import time
import json
import hashlib
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

JST_APP_KEY = "d561deb348274f1ba3505ec4578870fd"
JST_APP_SECRET = "84ad2c023b9b49378b1161ea569e383c"
JST_TOKEN = "cfda23ff97664494bc6fc5ab46f8ea48"
API_URL = "https://open.erp321.com/api/open/query.aspx"

def generate_sign(method, token, ts):
    params = {"token": token, "ts": ts}
    param_str = "".join(str(k) + str(v) for k, v in sorted(params.items()))
    sign_str = method + JST_APP_KEY + param_str + JST_APP_SECRET
    return hashlib.md5(sign_str.encode("utf-8")).hexdigest().lower()

def query_orders_page(shop_id, modified_begin, modified_end, page_index, page_size=100):
    ts = str(int(time.time()))
    method = "orders.single.query"
    sign = generate_sign(method, JST_TOKEN, ts)
    
    url = f"{API_URL}?method={method}&partnerid={JST_APP_KEY}&token={JST_TOKEN}&ts={ts}&sign={sign}"
    
    query_params = {
        "shop_id": str(shop_id),
        "modified_begin": modified_begin,
        "modified_end": modified_end,
        "page_index": page_index,
        "page_size": page_size
    }
    
    headers = {"Content-Type": "application/json; charset=utf-8"}
    response = requests.post(url, data=json.dumps(query_params, ensure_ascii=False).encode('utf-8'), headers=headers, timeout=30)
    return response.json()

def query_all_orders(shop_id, start_date, end_date):
    """Query all orders in date range, handling pagination"""
    all_orders = []
    current_start = start_date
    
    while current_start <= end_date:
        current_end = current_start + dt.timedelta(days=6)
        if current_end > end_date:
            current_end = end_date
        
        page = 1
        while True:
            result = query_orders_page(
                shop_id,
                current_start.strftime("%Y-%m-%d 00:00:00"),
                current_end.strftime("%Y-%m-%d 23:59:59"),
                page
            )
            
            if result.get("code") != 0:
                print(f"Error: {result}")
                break
            
            orders = result.get("orders", [])
            if not orders:
                break
            
            all_orders.extend(orders)
            if len(orders) < 100:
                break
            page += 1
            time.sleep(0.3)
        
        current_start = current_end + dt.timedelta(days=1)
        time.sleep(0.3)
    
    return all_orders

def export_orders_to_excel(orders, shop_name, start_date, end_date, output_path):
    """Export orders to Excel file matching the required format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "订单号", "收货人", "省", "市", "区", "收货地址", "电话号码",
        "商品名称", "数量", "单价", "小计",
        "国际运单号", "国内运单号", "查询链接",
        "合计", "订单日期", "备注"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row = 2
    
    for order in orders:
        # Get order-level info
        so_id = order.get("so_id", "")
        receiver_name = order.get("receiver_name", "") or ""
        receiver_state = (order.get("receiver_state", "") or "").replace("省", "").replace("市", "")
        receiver_city = (order.get("receiver_city", "") or "").replace("市", "")
        receiver_district = order.get("receiver_district", "") or ""
        receiver_address = order.get("receiver_address", "") or ""
        receiver_mobile = order.get("receiver_mobile", "") or ""
        receiver_phone = order.get("receiver_phone", "") or ""
        phone = receiver_mobile or receiver_phone
        
        order_date = (order.get("order_date", "") or "")[:10]
        pay_amount = float(order.get("pay_amount", 0) or 0)
        l_id = order.get("l_id", "") or ""  # 国内运单号
        cb_l_id = order.get("cb_l_id", "") or ""  # 国际运单号
        o_id = order.get("o_id", "")
        
        # Get items
        items = order.get("items", []) or []
        
        if items:
            # Each item becomes a row
            for item_idx, item in enumerate(items):
                # Item-level fields
                # Use 'name' field from the item, fallback to 'sku_name' or 'skus' field
                item_name = item.get("name") or item.get("sku_name") or order.get("skus", "") or ""
                qty = int(item.get("qty", 1) or 1)
                price = float(item.get("price", 0) or 0)
                subtotal = qty * price
                
                # Build row data - fill order info for ALL items (not just first)
                row_data = [
                    so_id,                                    # 订单号
                    receiver_name,                            # 收货人
                    receiver_state,                           # 省
                    receiver_city,                            # 市
                    receiver_district,                        # 区
                    receiver_address,                         # 收货地址
                    phone,                                    # 电话号码
                    item_name,                                # 商品名称
                    qty,                                      # 数量
                    price,                                    # 单价
                    subtotal,                                 # 小计
                    cb_l_id,                                 # 国际运单号
                    l_id,                                    # 国内运单号
                    f"https://erp.jushuitan.com/order/detail?o_id={o_id}", # 查询链接
                    pay_amount,                              # 合计
                    order_date,                              # 订单日期
                    ""                                        # 备注
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = thin_border
                    if col in [9, 10, 11, 15]:  # 数量、单价、小计、合计 - numeric
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = left_align
                
                current_row += 1
        else:
            # No items, write single row with skus field
            skus_str = order.get("skus", "") or ""
            row_data = [
                so_id, receiver_name, receiver_state, receiver_city,
                receiver_district, receiver_address, phone,
                skus_str, "", "", "",  # 商品名称, 数量, 单价, 小计
                cb_l_id, l_id,
                f"https://erp.jushuitan.com/order/detail?o_id={o_id}",
                pay_amount, order_date, ""
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
                if col in [9, 10, 11, 15]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = left_align
            
            current_row += 1
    
    # Adjust column widths
    col_widths = {
        1: 18,   # 订单号
        2: 10,   # 收货人
        3: 8,    # 省
        4: 8,    # 市
        5: 10,   # 区
        6: 35,   # 收货地址
        7: 15,   # 电话号码
        8: 35,   # 商品名称
        9: 6,    # 数量
        10: 10,  # 单价
        11: 10,  # 小计
        12: 22,  # 国际运单号
        13: 18,  # 国内运单号
        14: 40,  # 查询链接
        15: 10,  # 合计
        16: 12,  # 订单日期
        17: 15   # 备注
    }
    
    from openpyxl.utils import get_column_letter
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.row_dimensions[1].height = 25
    
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")
    return output_path

def main():
    # Configuration
    shop_id = 16871568  # A高总
    shop_name = "A高总"
    end_date = dt.datetime.now()
    start_date = end_date - dt.timedelta(days=7)
    
    print(f"Querying {shop_name}({shop_id}) from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    
    # Query orders
    all_orders = query_all_orders(shop_id, start_date, end_date)
    print(f"Total orders: {len(all_orders)}")
    
    if not all_orders:
        print("No orders found!")
        return
    
    # Count items
    total_items = sum(len(o.get("items", []) or []) for o in all_orders)
    print(f"Total line items: {total_items}")
    
    # Show sample
    sample = all_orders[0]
    items = sample.get("items", []) or []
    print(f"\nSample order: {sample.get('so_id')}")
    print(f"  Items count: {len(items)}")
    if items:
        print(f"  First item: {json.dumps(items[0], ensure_ascii=False)[:200]}")
    
    # Generate filename
    output_dir = "/opt/data/workspace/ustar-deploy/exports"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(
        output_dir,
        f"{shop_name}-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}-订单明细.xlsx"
    )
    
    # Export to Excel
    export_orders_to_excel(all_orders, shop_name, start_date, end_date, output_path)
    print(f"\nExport complete!")
    print(f"Output: {output_path}")
    
    # Verify output
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb.active
    print(f"Excel rows (including header): {ws.max_row}")
    print(f"Headers: {[cell.value for cell in ws[1]]}")
    print(f"Row 2: {[cell.value for cell in ws[2]]}")

if __name__ == "__main__":
    main()
